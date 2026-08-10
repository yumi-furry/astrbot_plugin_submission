import os
import re
import time
import uuid
import asyncio
import datetime
from typing import Optional

from astrbot.api.event import filter, AstrMessageEvent
from astrbot.api.star import Context, Star, register
from astrbot.api import AstrBotConfig, logger

import astrbot.api.message_components as Comp
from astrbot.core.utils.session_waiter import (
    session_waiter,
    SessionController,
)

from openpyxl import Workbook, load_workbook


# ============= 常量配置 =============

# 用户投稿状态（用于会话内状态机）
STATE_ASK_SUBMIT = "ask_submit"
STATE_ASK_CATEGORY = "ask_category"
STATE_UPLOAD_FILE = "upload_file"
STATE_ASK_TITLE = "ask_title"
STATE_ASK_COLLABORATOR = "ask_collaborator"
STATE_ASK_CONTACT = "ask_contact"

# 状态步骤顺序（用于回退上一步；第一个步骤无前置步骤）
STATE_ORDER = [
    STATE_ASK_SUBMIT,
    STATE_ASK_CATEGORY,
    STATE_UPLOAD_FILE,
    STATE_ASK_TITLE,
    STATE_ASK_COLLABORATOR,
    STATE_ASK_CONTACT,
]

# 直接退出关键词
EXIT_KEYWORDS = {"exit", "退出", "取消", "结束", "q", "quit", "stop"}
# 回退上一步关键词
BACK_KEYWORDS = {"back", "上一步", "返回", "后退", "prev", "previous", "撤销", "返回上一步"}

# 每一步提示中追加的快捷键说明（简洁版）
HINT_BACK_EXIT = "（输入 上一步 可回退 / 退出 可结束）"

# 支持的投稿类别
VALID_CATEGORIES = {"文本", "视频", "音频"}

# 各类别支持的文件扩展名（小写）
CATEGORY_EXTENSIONS = {
    "文本": {".txt", ".docx", ".doc", ".pdf", ".md", ".rtf"},
    "视频": {".mp4", ".avi", ".mov", ".mkv", ".flv", ".wmv", ".webm"},
    "音频": {".mp3", ".wav", ".flac", ".aac", ".ogg", ".m4a", ".wma"},
}

# 所有支持的扩展名合集（兜底提示用）
ALL_VALID_EXTENSIONS = set()
for _exts in CATEGORY_EXTENSIONS.values():
    ALL_VALID_EXTENSIONS.update(_exts)

# 每个状态执行成功后会写入哪些变量（用于回退时精准清空）
STATE_WRITTEN_VARS = {
    STATE_ASK_SUBMIT: [],
    STATE_ASK_CATEGORY: ["category"],
    STATE_UPLOAD_FILE: ["file_data", "file_name", "file_ext", "file_size"],
    STATE_ASK_TITLE: ["title"],
    STATE_ASK_COLLABORATOR: ["collaborator", "contact"],
    STATE_ASK_CONTACT: ["contact"],
}

# 作品名称正则过滤
TITLE_PATTERN = re.compile(r"^[\u4e00-\u9fa5a-zA-Z0-9\s_\-.,，。、]+$")

# 文件名非法字符替换
INVALID_FILENAME_CHARS = re.compile(r'[\\/:*?"<>|]')

# Excel表头
EXCEL_HEADERS = [
    "投稿ID", "用户ID", "用户昵称", "类别", "文件名",
    "作品名称", "联系方式", "投稿时间", "状态", "文件大小(KB)"
]


@register(
    "astrbot_plugin_submission",
    "AstrBot",
    "毛茸茸的跨年夜活动投稿审核插件",
    "1.0.0"
)
class SubmissionPlugin(Star):
    def __init__(self, context: Context, config: AstrBotConfig):
        super().__init__(context)
        # 保存配置引用（AstrBotConfig 继承自 dict，支持所有 dict 方法）
        self.config = config

        # 插件根目录
        self.plugin_dir = os.path.dirname(os.path.abspath(__file__))

        # 从配置读取关键参数（读取不到时使用默认值，兼容 _conf_schema 未初始化）
        def _cfg(key, default):
            val = config.get(key, default)
            return default if val is None else val

        self.data_dir = os.path.join(self.plugin_dir, "data")
        self.excel_path = os.path.join(self.data_dir, "submissions.xlsx")

        # 投稿文件保存目录（默认 data/files）
        files_dir_cfg = _cfg("files_dir", "data/files")
        if os.path.isabs(files_dir_cfg):
            self.files_dir = files_dir_cfg
        else:
            self.files_dir = os.path.join(self.plugin_dir, files_dir_cfg)

        # 存储限制（单位转换）
        self.max_total_storage_bytes = int(float(_cfg("max_total_storage_gb", 30.0)) * (1024 ** 3))
        self.max_single_file_bytes = int(_cfg("max_single_file_mb", 20)) * (1024 ** 2)

        # 业务限制
        self.daily_submit_limit = int(_cfg("daily_submit_limit", 3))
        self.session_timeout_sec = int(_cfg("session_timeout_sec", 60))
        self.only_private_mode = bool(_cfg("only_private_mode", False))

    async def initialize(self):
        """插件初始化：创建目录、Excel"""
        # 1. 数据目录
        os.makedirs(self.data_dir, exist_ok=True)
        logger.info(f"[投稿插件] 数据目录: {self.data_dir}")

        # 2. 文件保存目录
        os.makedirs(self.files_dir, exist_ok=True)
        logger.info(f"[投稿插件] 文件保存目录: {self.files_dir}")

        # 3. 初始化Excel
        self._init_excel()
        logger.info(f"[投稿插件] Excel存储: {self.excel_path}")

        # 4. 打印当前存储状态
        used = self._calc_dir_used_bytes(self.files_dir)
        used_gb = used / (1024 ** 3)
        limit_gb = self.max_total_storage_bytes / (1024 ** 3)
        logger.info(
            f"[投稿插件] 已用存储: {used_gb:.2f} GB / 限制 {limit_gb:.2f} GB"
        )

    async def terminate(self):
        """插件销毁"""
        logger.info("[投稿插件] 插件已卸载")

    # =========================================
    # 初始化辅助
    # =========================================

    def _init_excel(self):
        if not os.path.exists(self.excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "投稿记录"
            ws.append(EXCEL_HEADERS)
            wb.save(self.excel_path)

    # =========================================
    # 存储容量
    # =========================================

    @staticmethod
    def _calc_dir_used_bytes(directory: str) -> int:
        """计算目录下所有文件的总大小（字节）"""
        total = 0
        if not os.path.isdir(directory):
            return 0
        try:
            for root, _, files in os.walk(directory):
                for f in files:
                    fp = os.path.join(root, f)
                    try:
                        total += os.path.getsize(fp)
                    except OSError:
                        pass
        except Exception:
            pass
        return total

    def _check_storage_available(self, new_file_bytes: int) -> tuple:
        """
        检查是否有足够空间保存新文件。
        返回: (bool, 错误描述 或 空字符串)
        """
        used = self._calc_dir_used_bytes(self.files_dir)
        remaining = self.max_total_storage_bytes - used
        if remaining <= 0:
            used_gb = used / (1024 ** 3)
            limit_gb = self.max_total_storage_bytes / (1024 ** 3)
            return False, f"服务器投稿存储空间已满（已用 {used_gb:.2f}GB / {limit_gb:.2f}GB），请联系管理员。"
        if new_file_bytes > remaining:
            need_gb = new_file_bytes / (1024 ** 3)
            remain_gb = remaining / (1024 ** 3)
            return False, f"服务器剩余投稿空间不足（剩余 {remain_gb:.2f}GB，本次需要 {need_gb:.2f}GB），请联系管理员。"
        return True, ""

    # =========================================
    # 本地文件保存
    # =========================================

    def _save_file_local(self, file_bytes: bytes, display_name: str, ext: str) -> tuple:
        """
        保存投稿文件到本地。
        display_name: 用于拼接的友好名称（昵称_标题），不含扩展名。
        返回: (是否成功, 保存后的文件名(相对路径) 或 错误描述, 文件大小)
        """
        # 确保目录存在
        os.makedirs(self.files_dir, exist_ok=True)

        # 检查存储空间
        ok, err = self._check_storage_available(len(file_bytes))
        if not ok:
            return False, err, 0

        # 构造文件名（带时间戳和随机串，避免重名覆盖）
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        rand_suffix = uuid.uuid4().hex[:6]
        safe_name = INVALID_FILENAME_CHARS.sub("_", display_name)
        # 文件名长度兜底，避免过长
        if len(safe_name) > 60:
            safe_name = safe_name[:60]
        filename = f"{safe_name}_{timestamp}_{rand_suffix}{ext}"
        save_path = os.path.join(self.files_dir, filename)

        try:
            with open(save_path, "wb") as f:
                f.write(file_bytes)
            # 写入成功后，返回相对插件目录的路径，便于后续迁移
            rel_path = os.path.relpath(save_path, self.plugin_dir)
            # 统一使用正斜杠
            rel_path = rel_path.replace(os.sep, "/")
            return True, rel_path, len(file_bytes)
        except Exception as e:
            logger.error(f"[投稿插件] 保存文件失败: {e}, 路径={save_path}")
            return False, f"保存文件失败: {e}", 0

    # =========================================
    # 限流：每日投稿次数
    # =========================================

    def _get_today_submit_count(self, user_id: str) -> int:
        if not os.path.exists(self.excel_path):
            return 0
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            today = datetime.date.today().isoformat()
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) < 8:
                    continue
                row_user_id = str(row[1]) if row[1] is not None else ""
                submit_time = str(row[7]) if row[7] is not None else ""
                if row_user_id == str(user_id) and submit_time[:10] == today:
                    count += 1
            return count
        except Exception as e:
            logger.error(f"[投稿插件] 统计今日投稿数失败: {e}")
            return 0

    # =========================================
    # Excel 写入
    # =========================================

    def _append_submission_row(self, row_data: list) -> int:
        wb = load_workbook(self.excel_path)
        ws = wb.active
        if ws.max_row <= 1:
            new_id = 1
        else:
            last_id = ws.cell(row=ws.max_row, column=1).value
            try:
                new_id = int(last_id) + 1
            except (TypeError, ValueError):
                new_id = ws.max_row
        full_row = [new_id] + row_data[1:]
        ws.append(full_row)
        wb.save(self.excel_path)
        return new_id

    # =========================================
    # 通用工具
    # =========================================

    @staticmethod
    def _sanitize_filename(text: str) -> str:
        if not text:
            return "unnamed"
        return INVALID_FILENAME_CHARS.sub("_", text)

    # =========================================
    # 从消息事件中提取文件信息
    # =========================================

    async def _extract_file_from_event(self, event: AstrMessageEvent) -> Optional[tuple]:
        """
        从消息链中提取第一个 File/Video/Record/Image 组件对应的文件数据。
        返回 (file_bytes, file_name, file_size) 或 None。

        不同平台适配器的消息组件属性名可能略有差异，此处按常见顺序探测：
        - 文件名：name / filename / file_name
        - 文件大小：size / file_size / length
        - 优先使用官方异步 API：await get_file()（避免同步下载警告）
        - 兜底探测属性：
            - 文件字节：file / data / content / raw / binary
            - 本地路径：path / file_path / local_path / localPath
            - 远程 URL：url / src / file_url / link / resource
        """
        try:
            import requests as _requests  # 局部导入，避免无需要时依赖
        except Exception:
            _requests = None  # type: ignore

        message_chain = event.get_messages()
        for comp in message_chain:
            comp_name = type(comp).__name__
            if comp_name not in ("File", "Video", "Record", "Image"):
                continue

            # --- 1. 文件名 ---
            file_name = None
            for attr in ("name", "filename", "file_name"):
                val = getattr(comp, attr, None)
                if isinstance(val, str) and val:
                    file_name = val
                    break
            if not file_name:
                ext_map = {"File": ".bin", "Video": ".mp4", "Record": ".mp3", "Image": ".png"}
                file_name = f"uploaded_{int(time.time())}{ext_map.get(comp_name, '.bin')}"

            # --- 2. 文件大小 ---
            file_size = 0
            for attr in ("size", "file_size", "length"):
                val = getattr(comp, attr, None)
                if isinstance(val, int) and val > 0:
                    file_size = val
                    break

            # --- 3. 文件字节 ---
            file_bytes: Optional[bytes] = None

            # (a-1) 官方推荐异步方式：await get_file()
            #       避免直接访问 <File>.file 导致的"同步等待下载"警告
            if hasattr(comp, "get_file") and callable(getattr(comp, "get_file")):
                try:
                    result = await comp.get_file()
                    if isinstance(result, (bytes, bytearray, memoryview)):
                        file_bytes = bytes(result)
                    elif isinstance(result, str) and os.path.isfile(result):
                        # 部分平台返回的是本地文件路径
                        try:
                            with open(result, "rb") as f:
                                file_bytes = f.read()
                        except Exception as e:
                            logger.warning(f"[投稿插件] 读取get_file返回的本地文件失败 {result}: {e}")
                except Exception as e:
                    logger.warning(f"[投稿插件] 调用comp.get_file()失败，将回退到其他方式: {e}")

            # (a-2) 直接 bytes（兜底）
            if file_bytes is None:
                for attr in ("data", "content", "raw", "binary", "file"):
                    val = getattr(comp, attr, None)
                    if isinstance(val, (bytes, bytearray, memoryview)):
                        file_bytes = bytes(val)
                        break

            # (b) 本地路径
            if file_bytes is None:
                for attr in ("path", "file_path", "local_path", "localPath"):
                    val = getattr(comp, attr, None)
                    if isinstance(val, str) and val and os.path.isfile(val):
                        try:
                            with open(val, "rb") as f:
                                file_bytes = f.read()
                            break
                        except Exception as e:
                            logger.warning(f"[投稿插件] 读取本地文件失败 {val}: {e}")

            # (c) 远程 URL
            if file_bytes is None and _requests is not None:
                url = None
                for attr in ("url", "src", "file_url", "link", "resource"):
                    val = getattr(comp, attr, None)
                    if isinstance(val, str) and val.startswith(("http://", "https://")):
                        url = val
                        break
                if url:
                    try:
                        resp = _requests.get(url, timeout=30)
                        resp.raise_for_status()
                        file_bytes = resp.content
                    except Exception as e:
                        logger.warning(f"[投稿插件] 从URL下载文件失败 {url}: {e}")

            if file_bytes is not None:
                if file_size <= 0:
                    file_size = len(file_bytes)
                return file_bytes, file_name, file_size

        return None

    # =========================================
    # 投稿入口指令 /submit
    # =========================================

    @filter.command("submit")
    async def submit_command(self, event: AstrMessageEvent):
        """投稿入口：用户发送 /submit 开始投稿流程"""
        user_id = event.get_sender_id()
        user_name = event.get_sender_name()

        # 仅私信模式检查（优先尽早返回，避免后续IO开销）
        if self.only_private_mode:
            group_id = ""
            # 方式1：优先使用 get_group_id() 方法（跨平台通用）
            if hasattr(event, "get_group_id") and callable(getattr(event, "get_group_id")):
                try:
                    gid = event.get_group_id()
                    if gid:
                        group_id = str(gid)
                except Exception:
                    pass
            # 方式2：兜底读取 message_obj.group_id 属性
            if not group_id and hasattr(event, "message_obj"):
                group_id = str(getattr(event.message_obj, "group_id", "") or "")
            if group_id:
                yield event.plain_result("投稿功能仅支持私信（私聊）使用，请通过私聊发送 /submit 投稿～")
                return

        # 每日限流
        today_count = self._get_today_submit_count(user_id)
        if today_count >= self.daily_submit_limit:
            yield event.plain_result(
                f"你今天已经投稿{self.daily_submit_limit}次啦，明天再来吧～"
            )
            return

        # 第一步：询问是否投稿
        yield event.plain_result("📮 请问你要投稿吗？（请回复 是 / 否）" + HINT_BACK_EXIT)

        @session_waiter(timeout=self.session_timeout_sec, record_history_chains=False)
        async def submit_waiter(controller: SessionController, ev: AstrMessageEvent):
            """会话控制器：处理多轮对话"""
            try:
                # 优先使用 SessionController 的 session_data（若框架支持）
                if not hasattr(controller, "_sub_state"):
                    # 初始化会话状态（作为 controller 属性，避免依赖框架版本差异）
                    controller._sub_state = STATE_ASK_SUBMIT
                    controller._sub_category = None
                    controller._sub_file_data = None
                    controller._sub_file_name = None
                    controller._sub_file_ext = None
                    controller._sub_file_size = 0
                    controller._sub_title = None
                    controller._sub_collaborator = None
                    controller._sub_contact = None
                    controller._sub_user_name = user_name

                def state():
                    return getattr(controller, "_sub_state", STATE_ASK_SUBMIT)

                def set_state(s: str):
                    controller._sub_state = s

                def gv(name: str, default=None):
                    return getattr(controller, f"_sub_{name}", default)

                def sv(name: str, val):
                    setattr(controller, f"_sub_{name}", val)

                message_str = (ev.message_str or "").strip()
                msg_lower = message_str.lower()

                # ========== 通用指令：退出 / 回退上一步 ==========
                # 注意：在 STATE_ASK_SUBMIT 之前，用户发送 /submit 后的第一问（是/否）也支持直接 退出
                if msg_lower in EXIT_KEYWORDS:
                    await ev.send(ev.plain_result("已退出投稿流程。"))
                    controller.stop()
                    return

                if msg_lower in BACK_KEYWORDS:
                    # 查找当前状态的上一步
                    cur = state()
                    try:
                        idx = STATE_ORDER.index(cur)
                    except ValueError:
                        idx = 0
                    if idx <= 0:
                        # 已经在第一步
                        await ev.send(ev.plain_result("已经在第一步啦，无法回退～" + HINT_BACK_EXIT))
                        controller.keep(timeout=self.session_timeout_sec, reset_timeout=True)
                        return
                    # 回退
                    prev_state = STATE_ORDER[idx - 1]
                    prev_index = idx - 1
                    set_state(prev_state)

                    # 回退后精准清空：凡是 STATE_ORDER 中 index > prev_index 的步骤写入的变量都清空
                    vars_to_clear: set = set()
                    for i in range(prev_index + 1, len(STATE_ORDER)):
                        s = STATE_ORDER[i]
                        for v in STATE_WRITTEN_VARS.get(s, []):
                            vars_to_clear.add(v)
                    for field in vars_to_clear:
                        sv(field, None)

                    # 根据上一步状态发送对应的提示
                    if prev_state == STATE_ASK_SUBMIT:
                        await ev.send(ev.plain_result("📮 已回退到第一步：请问你要投稿吗？（请回复 是 / 否）" + HINT_BACK_EXIT))
                    elif prev_state == STATE_ASK_CATEGORY:
                        await ev.send(ev.plain_result("已回退到上一步。请选择投稿类别：文本 / 视频 / 音频" + HINT_BACK_EXIT))
                    elif prev_state == STATE_UPLOAD_FILE:
                        category = gv("category") or ""
                        if category and category in CATEGORY_EXTENSIONS:
                            ext_list = ", ".join(sorted(CATEGORY_EXTENSIONS[category]))
                        else:
                            ext_list = ", ".join(sorted(ALL_VALID_EXTENSIONS))
                        await ev.send(ev.plain_result(f"已回退到上一步。请重新上传你的作品文件（支持 {ext_list}）" + HINT_BACK_EXIT))
                    elif prev_state == STATE_ASK_TITLE:
                        await ev.send(ev.plain_result("已回退到上一步。请为作品重新起个名字吧" + HINT_BACK_EXIT))
                    elif prev_state == STATE_ASK_COLLABORATOR:
                        await ev.send(ev.plain_result("已回退到上一步。请问是否有其他创作者合作？（请回复 是 / 否）" + HINT_BACK_EXIT))
                    controller.keep(timeout=self.session_timeout_sec, reset_timeout=True)
                    return

                # ========== 状态机 ==========
                if state() == STATE_ASK_SUBMIT:
                    if message_str == "是":
                        set_state(STATE_ASK_CATEGORY)
                        await ev.send(ev.plain_result("请选择投稿类别：文本 / 视频 / 音频" + HINT_BACK_EXIT))
                        controller.keep(timeout=self.session_timeout_sec, reset_timeout=True)
                    elif message_str == "否":
                        await ev.send(ev.plain_result("好的，有需要再来找我～"))
                        controller.stop()
                    else:
                        await ev.send(ev.plain_result("📮 请问你要投稿吗？（请回复 是 / 否）" + HINT_BACK_EXIT))
                        controller.keep(timeout=self.session_timeout_sec, reset_timeout=True)
                    return

                if state() == STATE_ASK_CATEGORY:
                    if message_str in VALID_CATEGORIES:
                        set_state(STATE_UPLOAD_FILE)
                        sv("category", message_str)
                        ext_list = ", ".join(sorted(CATEGORY_EXTENSIONS[message_str]))
                        await ev.send(ev.plain_result(f"请上传你的作品文件（支持 {ext_list}）" + HINT_BACK_EXIT))
                        controller.keep(timeout=self.session_timeout_sec, reset_timeout=True)
                    else:
                        await ev.send(ev.plain_result("暂不支持该类别，请选择：文本 / 视频 / 音频" + HINT_BACK_EXIT))
                        controller.keep(timeout=self.session_timeout_sec, reset_timeout=True)
                    return

                if state() == STATE_UPLOAD_FILE:
                    file_info = await self._extract_file_from_event(ev)
                    if file_info is None:
                        await ev.send(ev.plain_result("不支持此格式，请重新上传" + HINT_BACK_EXIT))
                        controller.keep(timeout=self.session_timeout_sec, reset_timeout=True)
                        return

                    file_bytes, file_name, file_size = file_info

                    # 大小检查
                    if file_size > self.max_single_file_bytes:
                        limit_mb = self.max_single_file_bytes // (1024 ** 2)
                        await ev.send(
                            ev.plain_result(
                                f"文件过大（{file_size // 1024}KB），单次最大{limit_mb}MB，"
                                f"请压缩后重新上传" + HINT_BACK_EXIT
                            )
                        )
                        controller.keep(timeout=self.session_timeout_sec, reset_timeout=True)
                        return

                    # 扩展名与类别匹配
                    ext = os.path.splitext(file_name)[1].lower()
                    category = gv("category")
                    allowed_exts = CATEGORY_EXTENSIONS.get(category, set())
                    if ext not in allowed_exts:
                        allowed = ", ".join(sorted(allowed_exts))
                        await ev.send(
                            ev.plain_result(
                                f"文件格式不符合{category}类别（允许: {allowed}），"
                                f"请重新上传" + HINT_BACK_EXIT
                            )
                        )
                        controller.keep(timeout=self.session_timeout_sec, reset_timeout=True)
                        return

                    sv("file_data", file_bytes)
                    sv("file_name", file_name)
                    sv("file_ext", ext)
                    sv("file_size", file_size)
                    set_state(STATE_ASK_TITLE)

                    logger.info(
                        f"[投稿插件] 用户 {user_id}({user_name}) 上传文件校验通过: "
                        f"类别={category}, 文件={file_name}, 大小={file_size}"
                    )
                    await ev.send(ev.plain_result("文件已收到！请为作品起个名字吧" + HINT_BACK_EXIT))
                    controller.keep(timeout=self.session_timeout_sec, reset_timeout=True)
                    return

                if state() == STATE_ASK_TITLE:
                    if not message_str:
                        await ev.send(ev.plain_result("作品名称不能为空，请重新输入～" + HINT_BACK_EXIT))
                        controller.keep(timeout=self.session_timeout_sec, reset_timeout=True)
                        return
                    if not TITLE_PATTERN.match(message_str):
                        await ev.send(ev.plain_result("内容包含非法字符，请重新输入～" + HINT_BACK_EXIT))
                        controller.keep(timeout=self.session_timeout_sec, reset_timeout=True)
                        return
                    sv("title", message_str)
                    set_state(STATE_ASK_COLLABORATOR)
                    await ev.send(ev.plain_result("请问是否有其他创作者合作？（请回复 是 / 否）" + HINT_BACK_EXIT))
                    controller.keep(timeout=self.session_timeout_sec, reset_timeout=True)
                    return

                if state() == STATE_ASK_COLLABORATOR:
                    if message_str == "是":
                        sv("collaborator", True)
                        set_state(STATE_ASK_CONTACT)
                        await ev.send(ev.plain_result("请输入合作者联系方式（QQ / 微信 / 手机号等均可）" + HINT_BACK_EXIT))
                        controller.keep(timeout=self.session_timeout_sec, reset_timeout=True)
                    elif message_str == "否":
                        sv("collaborator", False)
                        sv("contact", "")
                        await self._do_save(ev, gv, user_id, user_name)
                        controller.stop()
                    else:
                        await ev.send(ev.plain_result("请问是否有其他创作者合作？（请回复 是 / 否）" + HINT_BACK_EXIT))
                        controller.keep(timeout=self.session_timeout_sec, reset_timeout=True)
                    return

                if state() == STATE_ASK_CONTACT:
                    # 不再限制格式，任意内容都可（包括 QQ、微信、手机号、邮箱等）
                    if not message_str:
                        await ev.send(ev.plain_result("联系方式不能为空，请重新输入" + HINT_BACK_EXIT))
                        controller.keep(timeout=self.session_timeout_sec, reset_timeout=True)
                        return
                    sv("contact", message_str)
                    await self._do_save(ev, gv, user_id, user_name)
                    controller.stop()
                    return

            except Exception as _e:
                logger.error(f"[投稿插件] 会话内部异常: {_e}")
                try:
                    await ev.send(ev.plain_result(f"发生错误：{str(_e)}，请联系管理员"))
                except Exception:
                    pass
                controller.stop()

        # 启动会话控制器
        try:
            await submit_waiter(event)
        except TimeoutError:
            logger.info(f"[投稿插件] 用户 {user_id} 投稿超时，已退出")
            yield event.plain_result("投稿超时，已退出")
        except Exception as e:
            logger.error(f"[投稿插件] 会话异常: {e}")
            yield event.plain_result(f"发生错误，请联系管理员: {str(e)}")
        finally:
            event.stop_event()

    # =========================================
    # 保存投稿流程
    # =========================================

    async def _do_save(self, ev: AstrMessageEvent, gv, user_id: str, user_name: str):
        """保存投稿：本地文件存储 + Excel写入"""
        category = gv("category")
        file_bytes: bytes = gv("file_data")
        file_ext: str = gv("file_ext")
        title: str = gv("title")
        contact: str = gv("contact") or ""
        file_size: int = gv("file_size") or len(file_bytes or b"")
        file_size_kb = file_size // 1024

        # 1. 构造显示文件名（不包含扩展名，用于保存函数拼接）
        safe_nick = self._sanitize_filename(user_name)
        safe_title = self._sanitize_filename(title)
        display_name = f"{safe_nick}_{safe_title}"

        # 2. 保存文件到本地（含容量检查）
        # 注意：写磁盘是同步IO，丢到线程池执行避免阻塞事件循环
        try:
            loop = asyncio.get_event_loop()
            ok, name_or_err, actual_size = await loop.run_in_executor(
                None, self._save_file_local, file_bytes, display_name, file_ext
            )
        except Exception as e:
            logger.error(f"[投稿插件] 文件保存异常: {e}")
            ok, name_or_err, actual_size = False, f"保存异常: {e}", 0

        if not ok:
            logger.error(f"[投稿插件] 用户 {user_id} 文件保存失败: {name_or_err}")
            await ev.send(ev.plain_result(f"保存失败：{name_or_err}，请稍后重试或联系管理员"))
            return

        saved_rel_path = name_or_err  # 保存后返回的相对路径
        logger.info(f"[投稿插件] 用户 {user_id} 文件保存成功: {saved_rel_path} ({actual_size} bytes)")

        # 3. 写入Excel
        submit_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        status = "待审核"
        row = [
            0,  # 投稿ID占位
            str(user_id),
            user_name,
            category,
            saved_rel_path,  # 文件名：使用相对路径，方便迁移
            title,
            contact,
            submit_time,
            status,
            file_size_kb,
        ]
        try:
            submission_id = self._append_submission_row(row)
            logger.info(
                f"[投稿插件] 投稿成功: 投稿ID={submission_id} 用户={user_id} "
                f"类别={category} 作品名={title}"
            )
        except Exception as e:
            logger.error(f"[投稿插件] 写入Excel失败: {e}")
            await ev.send(ev.plain_result("保存投稿记录失败，请联系管理员"))
            return

        # 4. 成功提示
        await ev.send(ev.plain_result("✅ 提交成功！我们将在5个工作日内完成审核～"))
